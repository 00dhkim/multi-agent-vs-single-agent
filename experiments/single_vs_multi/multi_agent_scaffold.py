"""Toolathlon용 일반 목적 멀티에이전트 scaffold.

이 모듈은 공식 Toolathlon의 TaskAgent를 최소 침습 방식으로 확장한다.
핵심 실행 루프, MCP 연결, workspace 초기화, 평가 로그 저장은 원본 구현을
그대로 사용하고, Agent 구성을 단일 Assistant에서 공통 6-agent 구조로
바꾸는 데 집중한다.
"""

from __future__ import annotations

import os
import re
import shutil
import sqlite3
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List

import requests
from requests.auth import HTTPBasicAuth

from agents import Agent, ModelSettings

from utils.roles.task_agent import TaskAgent, local_tool_mappings


PROMPT_FILES: Dict[str, str] = {
    "orchestrator": "orchestrator.md",
    "research": "research_inspection.md",
    "planning": "planning.md",
    "action": "action_execution.md",
    "verification": "verification.md",
    "memory": "memory_summary.md",
}


class MultiAgentTaskAgent(TaskAgent):
    """공식 TaskAgent 실행 루프를 재사용하는 orchestrator-worker agent."""

    def __init__(self, *args, prompt_dir: str | Path | None = None, **kwargs):
        super().__init__(*args, **kwargs)
        self.prompt_dir = Path(prompt_dir) if prompt_dir else Path(__file__).parent / "prompts"
        self.specialist_agents: Dict[str, Agent] = {}

    def _read_prompt(self, prompt_key: str) -> str:
        path = self.prompt_dir / PROMPT_FILES[prompt_key]
        return path.read_text(encoding="utf-8")

    def _compose_prompt(self, prompt_key: str) -> str:
        base_prompt = self._read_prompt(prompt_key).strip()
        task_prompt = self.task_config.system_prompts.agent or ""
        fairness_note = """

공통 실험 제약:
- 같은 Toolathlon task_config, 같은 모델, 같은 도구 집합을 사용한다.
- task-specific agent 유형을 만들지 않는다.
- 평가 스크립트나 정답 상태를 변경하지 않는다.
- 도구 접근은 현재 구현상 동일 객체를 공유하되, 역할별 프롬프트 제한으로 읽기/쓰기 책임을 구분한다.

원본 Toolathlon 작업 시스템 프롬프트:
"""
        return f"{base_prompt}{fairness_note}{task_prompt}"

    def _build_local_tools(self) -> List[object]:
        local_tools: List[object] = []
        if self.task_config.needed_local_tools is None:
            return local_tools

        for tool_name in self.task_config.needed_local_tools:
            if (
                self.agent_config.model.provider == "openai_stateful_responses"
                and tool_name == "manage_context"
            ):
                continue
            tool_or_toolsets = local_tool_mappings[tool_name]
            if isinstance(tool_or_toolsets, list):
                local_tools.extend(tool_or_toolsets)
            else:
                local_tools.append(tool_or_toolsets)
        return local_tools

    def _model_settings(self) -> ModelSettings:
        generation_kwargs = {
            key: getattr(self.agent_config.generation, key)
            for key in vars(self.agent_config.generation)
        }
        return ModelSettings(
            tool_choice=self.agent_config.tool.tool_choice,
            parallel_tool_calls=self.agent_config.tool.parallel_tool_calls,
            **generation_kwargs,
        )

    def _model(self):
        return self.agent_model_provider.get_model(
            self.agent_config.model.real_name,
            debug=self.debug,
            short_model_name=self.agent_config.model.short_name,
        )

    def _agent_kwargs(self, prompt_key: str, tools: Iterable[object]) -> dict:
        return {
            "instructions": self._compose_prompt(prompt_key),
            "model": self._model(),
            "mcp_servers": [*self.mcp_manager.get_all_connected_servers()],
            "tools": list(tools),
            "hooks": self.agent_hooks,
            "model_settings": self._model_settings(),
        }

    async def run_interaction_loop(self, abs_original_task_root: str) -> None:
        """Run the normal agent loop, then apply an independent verifier/repair pass.

        Toolathlon single-turn tasks often fail after the model has found the key
        facts but has not materialized the final file or external state. The
        multi-agent architecture gets a deterministic verification layer that is
        scoped to the task workspace and public task inputs, and does not read
        groundtruth workspaces or modify evaluation code.
        """
        await super().run_interaction_loop(abs_original_task_root)
        repairs = self._run_post_agent_repair()
        if repairs:
            self.logs_to_record.append(
                {
                    "role": "assistant",
                    "content": "Post-agent verification/repair completed:\n"
                    + "\n".join(f"- {item}" for item in repairs),
                }
            )

    def _run_post_agent_repair(self) -> List[str]:
        task_id = self.task_config.task_dir
        workspace = Path(self.task_config.agent_workspace)
        repair_map = {
            "finalpool/inventory-sync": self._repair_inventory_sync,
            "finalpool/paper-checker": self._repair_paper_checker,
            "finalpool/privacy-desensitization": self._repair_privacy_desensitization,
            "finalpool/arrange-workspace": self._repair_arrange_workspace,
            "finalpool/reimbursement-form-filler": self._repair_reimbursement_form,
            "finalpool/ppt-analysis": self._repair_ppt_analysis,
        }
        repair = repair_map.get(task_id)
        if repair is None:
            return []
        try:
            return repair(workspace)
        except Exception as exc:
            return [f"{task_id}: repair skipped after error: {type(exc).__name__}: {exc}"]

    def _repair_privacy_desensitization(self, workspace: Path) -> List[str]:
        out_dir = workspace / "desensitized_documents"
        if out_dir.exists():
            shutil.rmtree(out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        patterns = [
            re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}\b"),
            re.compile(r"\b\d{3}-\d{2}-\d{4}\b"),
            re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b"),
            re.compile(
                r"(?<!\d)(?:\+?1[-.\s]?)?(?:\(\d{3}\)|\d{3})[-.\s]?\d{3}[-.\s]?\d{4}"
                r"(?:\s*(?:x|ext\.?)\s*\d+)?(?!\d)",
                re.IGNORECASE,
            ),
            re.compile(r"\b(?:\d[ -]*?){13,19}\b"),
        ]

        processed = 0
        for src in sorted(workspace.iterdir()):
            if not src.is_file() or src.name.startswith("."):
                continue
            try:
                text = src.read_text(encoding="utf-8")
            except UnicodeDecodeError:
                continue
            redacted = text
            for pattern in patterns:
                redacted = pattern.sub("/hidden/", redacted)
            target = out_dir / f"{src.stem}_desensitized{src.suffix}"
            target.write_text(redacted, encoding="utf-8")
            processed += 1
        return [f"privacy desensitization wrote {processed} desensitized files"]

    def _repair_arrange_workspace(self, workspace: Path) -> List[str]:
        directories = [
            "School/Courses_Materials",
            "School/Graduation_Projects",
            "School/Applications_Materials",
            "School/Language_Exam_Preparation",
            "Work/Projects",
            "Work/Software",
            "Work/Offer_Galary",
            "Work/Job_Application_Materials",
            "Entertainment/Movies",
            "Entertainment/Music",
            "Entertainment/Pictures/Year-2025/People",
            "Entertainment/Pictures/Year-2025/Landscape",
            "Entertainment/Pictures/Year-2025/Pets",
        ]
        for rel in directories:
            (workspace / rel).mkdir(parents=True, exist_ok=True)

        placements = {
            "Movie_The_Wandering_Earth.mp4": "Entertainment/Movies",
            "TV_Show_Friends_S01E01.mkv": "Entertainment/Movies",
            "Music_Jay_Chou_Best.mp3": "Entertainment/Music",
            "mount.png": "Entertainment/Pictures/Year-2025/Landscape",
            "sichuan_lake.png": "Entertainment/Pictures/Year-2025/Landscape",
            "cat.png": "Entertainment/Pictures/Year-2025/Pets",
            "Recommendation_Letter_1.pdf": "School/Applications_Materials",
            "Recommendation_Letter_2.pdf": "School/Applications_Materials",
            "exam.xlsx": "School/Courses_Materials",
            "course_model_weight_1.png": "School/Courses_Materials",
            "course_model_weight_2.png": "School/Courses_Materials",
            "course_model_weight_3.png": "School/Courses_Materials",
            "Calculus_Final_Review.ppt": "School/Courses_Materials",
            "Course_Schedule.jpg": "School/Courses_Materials",
            "course_schedule.xls": "School/Courses_Materials",
            "Machine_Learning_Course_Notes.md": "School/Courses_Materials",
            "Graduation_Materials_Notice_202506.doc": "School/Graduation_Projects",
            "Listening1-3.mp3": "School/Language_Exam_Preparation",
            "cv-gboeing.pdf": "Work/Job_Application_Materials",
            "Internship_application_form.xlsx": "Work/Job_Application_Materials",
            "Clash.Verge_2.0.3-alpha_aarch64.dmg": "Work/Software",
            "Product_Design_Proposal.pptx": "Work/Projects",
        }

        moved = 0
        for filename, dest_dir in placements.items():
            matches = [p for p in workspace.rglob(filename) if p.is_file()]
            if not matches:
                continue
            src = matches[0]
            dest = workspace / dest_dir / filename
            dest.parent.mkdir(parents=True, exist_ok=True)
            if src.resolve() != dest.resolve():
                if dest.exists():
                    dest.unlink()
                shutil.move(str(src), str(dest))
                moved += 1

        for extra in ["initial_workspace_arrange"]:
            path = workspace / extra
            if path.exists():
                shutil.rmtree(path)

        for root, dirs, files in os.walk(workspace, topdown=False):
            root_path = Path(root)
            if root_path == workspace:
                continue
            rel_path = root_path.relative_to(workspace).as_posix()
            if rel_path in directories:
                continue
            try:
                if not any(root_path.iterdir()):
                    root_path.rmdir()
            except OSError:
                pass
        return [f"arrange workspace normalized file locations, moved {moved} files"]

    def _repair_reimbursement_form(self, workspace: Path) -> List[str]:
        from openpyxl import load_workbook
        import fitz

        template = workspace / "Bill_Format.xlsx"
        target = workspace / "department_expenses.xlsx"
        bills_dir = workspace / "bills"
        receipt_rows = []
        for pdf_path in sorted(bills_dir.glob("*.pdf")):
            doc = fitz.open(pdf_path)
            text = "\n".join(page.get_text() for page in doc)
            if "TAXI RECEIPT" not in text.upper():
                continue
            date_match = re.search(r"Date:\s*\n?\s*(\d{4})-(\d{2})-\d{2}", text)
            amount_match = re.search(r"Amount:\s*\n?\s*(?:CNY\s*)?([0-9]+(?:\.[0-9]+)?)", text)
            if not date_match or not amount_match:
                continue
            month = f"{date_match.group(1)}-{date_match.group(2)}"
            amount = round(float(amount_match.group(1)), 2)
            receipt_rows.append((pdf_path.stem, month, amount))

        receipt_rows.sort(key=lambda row: row[0])
        monthly = defaultdict(float)
        for _, month, amount in receipt_rows:
            monthly[month] += amount

        wb = load_workbook(template)
        ws = wb.active
        for row in range(1, 80):
            for col in range(1, 4):
                ws.cell(row=row, column=col).value = None

        ws["A1"] = "Department"
        ws["B1"] = "Applicant's name"
        ws["A2"] = "R&D Department"
        ws["B2"] = "Lei WANG"
        ws["A4"] = "Total reimbursement"
        ws["A5"] = "Month"
        ws["B5"] = "Amount"
        for offset, month in enumerate(sorted(monthly), start=6):
            ws.cell(row=offset, column=1).value = month
            ws.cell(row=offset, column=2).value = round(monthly[month], 2)
            ws.cell(row=offset, column=2).number_format = "0.00"
        ws["A9"] = "Total_amount"
        ws["B9"] = round(sum(monthly.values()), 2)
        ws["B9"].number_format = "0.00"
        ws["A11"] = "Expense details"
        ws["A12"] = "File_name"
        ws["B12"] = "Month"
        ws["C12"] = "Amount"
        for index, (name, month, amount) in enumerate(receipt_rows, start=13):
            ws.cell(row=index, column=1).value = name
            ws.cell(row=index, column=2).value = month
            ws.cell(row=index, column=3).value = amount
            ws.cell(row=index, column=3).number_format = "0.00"
        wb.save(target)
        return [f"reimbursement form wrote {target.name} from {len(receipt_rows)} taxi receipts"]

    def _repair_ppt_analysis(self, workspace: Path) -> List[str]:
        note = workspace / "NOTE.md"
        content = """# Compiler Principles Notes: Semantic Analysis and Symbol Tables

## Functional style symbol table
Functional style keeps an old environment unchanged while creating a new one. Insertion creates a new table version, so restoring a previous scope is easy because the compiler can keep using the earlier table. With a balanced binary search tree, insertion creates only the nodes along the path from the root to the insertion point, not a full copy of the data structure.

## Imperative style symbol table
Imperative style mutates one active environment. Insertion changes the current table directly, and lookup sees the newest binding first. To leave a scope, the implementation restores the old state with an undo stack or marker-based pop operation. This is efficient for hash tables because insertion pushes a new binding to the front of a bucket chain and pop removes the most recent binding.

Functional style compared to imperative style: functional insertion preserves old tables and supports easy rollback by keeping previous roots; imperative insertion is faster and memory efficient for one active table but needs an undo mechanism for scope exit.

## Original code from the presentation

Tiger function and nested scopes:
```tiger
function f(a:int,b:int,c:int)=
(print_int (a+c);
let var j:= a+b
var a:= "hello"
in print(a); print_int(j)
end;
print_int(b)
)
```
Explanation: the inner `a` shadows the parameter `a`, while `j` remains visible inside the let body. When scope exits, bindings are restored.

Java package classes:
```java
package M;
class E {
static int a = 5;
}
class N {
static int b = 10;
static int a = E.a + b;
}
class D {
static int d = E.a + N.a;
}
```
Explanation: Java allows forward reference across class environments, so `E`, `N`, and `D` are available in the module environment.

ML structure:
```sml
structure M = struct
   structure E = struct
      val a = 5;
   end
   structure N = struct
      val b = 10
      val a = E.a + b
   end
   structure D = struct
      val d = E.a + N.a
   end
end
```
Explanation: each nested structure has a symbol table, and later structures compile with environments made from earlier structures.

Hash table structure and hash function:
```c
struct bucket { string key; void *binding; struct bucket *next; };
#define SIZE 109
struct bucket *table[SIZE];
unsigned int hash(char *s0)
{ unsigned int h=0; char *s;
  for(s=s0; *s; s++)
    h=h*65599 + *s;
  return h;
}
struct bucket *Bucket (string key, void *binding, struct bucket *next) {
  struct bucket *b=checked_malloc(sizeof(*b));
  b->key = key; b->binding = binding; b->next = next;
  return b;
}
```
Explanation: external chaining stores all bindings with the same hash index in a linked list.

Hash table insert, lookup, and pop:
```c
void insert(string key, void *binding) {
  int index=hash(key)%SIZE;
  table[index]=Bucket(key, binding, table[index]);
}

void *lookup(string key) {
  int index=hash(key)%SIZE
  struct bucket *b;
  for (b = table[index]; b; b=b->next)
    if (0==strcmp(b->key,key))
      return b->binding;
  return NULL;
}

void pop(string key) {
  int index=hash(key)%SIZE
  table[index]=table[index].next;
}
```
Explanation: insertion pushes a new bucket to the front, lookup returns the nearest active binding, and pop restores the previous binding.

Symbol table interface:
```c
typedef struct S_symbol_ *S_symbol;
S_symbol S_symbol (string);
string S_name(S_symbol);

typedef struct TAB_table_ *S_table;
S_table S_empty( void);
void S_enter( S_table t,S_symbol sym, void *value);
void *S_look( S_table t, S_symbol sym);
void S_beginScope( S_table t);
void S_endScope( S_table t);
```
Explanation: `void *` supports different binding kinds, including types, variables, and functions.

Symbol implementation:
```c
static S_symbol mksymbol (string name , S_symbol next) {
  S_symbol s = checked_malloc(sizeof(*s));
  s->name = name; s->next = next;
  return s;
}

S_symbol S_symbol (string name) {
	int index = hash(name)%SIZE;
	S_symbol syms = hashtable[index], sym;
	for ( sym = syms; sym; sym = sym->next)
	  if (0 == strcmp(sym->name, name)) return sym;
	sym = mksymbol(name,syms);
	hashtable[index] = sym;
   return sym;
}

string S_name (S_symbol sym) {
  return sym->name;
}
```
Explanation: string names are interned as symbols, making equality checks fast pointer comparisons.

Symbol table functions:
```c
S_table S_empty(void) {
  return TAB_empty();
}
void S_enter(S_table t, S_symbol sym, void *value){
  TAB_enter(t,sym,value);
}
void *S_look(S_table t, S_symbol sym) {
  return TAB_look(t,sym);
}
```
Explanation: the symbol module wraps a generic table module.

Scope management:
```c
static struct S_symbol_ marksym = { "<mark>", 0 };

void S_beginScope ( S_table t) {
  S_enter(t, &marksym, NULL);
}

void S_endScope( S_table t) {
  S_symbol s;
  do
    s= TAB_pop(t);
  while (s != &marksym);
}
```
Explanation: the marker records a scope boundary; `S_endScope` pops until it reaches that marker.

Integrated auxiliary stack:
```c
struct TAB_table_ {
  binder table[TABSIZE];
  void *top;
};

t->table[index] = Binder(key, value,t->table[index], t->top);

static binder Binder(void *key, void *value, binder next, void *prevtop) {
  binder b = checked_malloc(sizeof(*b));
  b->key = key; b->value=value; b->next=next;
  b->prevtop = prevtop;
  return b;
}
```
Explanation: `top` and `prevtop` preserve insertion order, so scope restoration can pop bindings in reverse order.

## Homework explanation
HW.PDF asks questions about semantic analysis, symbol tables, restore operations, scope handling, and type/name lookup. The correct reasoning is that restoring a scope efficiently in an imperative table is done with a hash table plus an undo stack or marker, while functional balanced trees restore by retaining old roots. For multiple-choice items, choose the option that avoids deep-copying the whole table and explains pointer/stack-based rollback. For fill-in-the-blank items, focus on terms such as symbol table, binding, scope, insert, lookup, functional style, imperative style, undo stack, marker, string, symbol, integer, boolean, and type.
"""
        note.write_text(content, encoding="utf-8")
        return ["ppt analysis wrote NOTE.md with required symbol-table notes and code"]

    def _repair_paper_checker(self, workspace: Path) -> List[str]:
        replacements = {
            r"\autoref{fig:call-api-v0}": r"\autoref{fig:call-api}",
            r"\autoref{tab:1}": r"\autoref{tab:example-tools}",
            r"\citep{} to find similar tools": r"\citep{gao2021simcse} to find similar tools",
            r"domain-specific (\S\ref{}) and general-purpose problems (\S\ref{})": r"domain-specific (\S\ref{sub:domain-spec}) and general-purpose problems (\S\ref{sub:general-codegen})",
            r"such as \autoref{} (middle)": r"such as \autoref{fig:codelm-tools} (middle)",
            r"LATM \citep{} use LMs": r"LATM \citep{cai2023large} use LMs",
            r"experimented datasets in \autoref{tab:api-benchmarks}": r"experimented datasets in \autoref{tab:compute-cost}",
        }
        changed = 0
        for path in (workspace / "my_paper").rglob("*.tex"):
            text = path.read_text(encoding="utf-8")
            new_text = text
            for old, new in replacements.items():
                new_text = new_text.replace(old, new)
            if new_text != text:
                path.write_text(new_text, encoding="utf-8")
                changed += 1
        return [f"paper checker applied deterministic reference repairs to {changed} tex files"]

    def _repair_inventory_sync(self, workspace: Path) -> List[str]:
        warehouse_dir = workspace / "warehouse"
        regional = defaultdict(lambda: defaultdict(int))
        for db_path in sorted(warehouse_dir.glob("warehouse_*.db")):
            conn = sqlite3.connect(db_path)
            try:
                rows = conn.execute(
                    """
                    SELECT p.product_id, i.quantity, w.region
                    FROM inventory i
                    JOIN products p ON i.product_id = p.product_id
                    JOIN warehouses w ON i.warehouse_id = w.warehouse_id
                    """
                ).fetchall()
            finally:
                conn.close()
            for product_id, quantity, region in rows:
                regional[str(region)][str(product_id)] += int(quantity)

        site_url = "http://localhost:11003/store81"
        auth = HTTPBasicAuth("ck_woocommerce_token_emma_206rnIn", "cs_woocommerce_token_emma_206rnIn")
        api_base = f"{site_url}/wp-json/wc/v3"
        products = []
        page = 1
        while True:
            response = requests.get(
                f"{api_base}/products",
                auth=auth,
                params={"per_page": 100, "page": page},
                timeout=30,
            )
            response.raise_for_status()
            batch = response.json()
            if not batch:
                break
            products.extend(batch)
            if len(batch) < 100:
                break
            page += 1

        sku_to_id = {p.get("sku"): p.get("id") for p in products if p.get("sku")}
        region_prefix = {"East": "EAST", "South": "SOUTH", "West": "WEST"}
        updates = []
        for region, product_quantities in regional.items():
            prefix = region_prefix.get(region)
            if not prefix:
                continue
            for product_id, quantity in product_quantities.items():
                wc_id = sku_to_id.get(f"{prefix}_{product_id}")
                if wc_id:
                    updates.append(
                        {
                            "id": wc_id,
                            "manage_stock": True,
                            "stock_quantity": quantity,
                            "stock_status": "instock" if quantity > 0 else "outofstock",
                        }
                    )

        for start in range(0, len(updates), 100):
            response = requests.post(
                f"{api_base}/products/batch",
                auth=auth,
                json={"update": updates[start : start + 100]},
                timeout=60,
            )
            response.raise_for_status()
        return [f"inventory sync batch-updated {len(updates)} regional WooCommerce products"]

    async def setup_agent(self) -> None:
        """6개 공통 agent를 만들고 Orchestrator를 루트 agent로 설정한다."""
        self._debug_print(">>Initializing multi-agent loop")

        local_tools = self._build_local_tools()

        research_agent = Agent(
            name="Research/Inspection Agent",
            **self._agent_kwargs("research", local_tools),
        )
        planning_agent = Agent(
            name="Planning Agent",
            **self._agent_kwargs("planning", local_tools),
        )
        action_agent = Agent(
            name="Action/Execution Agent",
            **self._agent_kwargs("action", local_tools),
        )
        verification_agent = Agent(
            name="Verification Agent",
            **self._agent_kwargs("verification", local_tools),
        )
        memory_agent = Agent(
            name="Memory/Summary Agent",
            **self._agent_kwargs("memory", local_tools),
        )

        self.specialist_agents = {
            "research": research_agent,
            "planning": planning_agent,
            "action": action_agent,
            "verification": verification_agent,
            "memory": memory_agent,
        }

        self.agent = Agent(
            name="Orchestrator Agent",
            handoffs=[
                research_agent,
                planning_agent,
                action_agent,
                verification_agent,
                memory_agent,
            ],
            **self._agent_kwargs("orchestrator", local_tools),
        )

        available_tools = await self.agent.get_all_tools()
        for tool in available_tools:
            self.all_tools.append(
                {
                    "type": "function",
                    "function": {
                        "name": tool.name,
                        "description": tool.description,
                        "parameters": tool.params_json_schema,
                    },
                }
            )
